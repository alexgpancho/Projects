# Importar librerías
import re
import glob
import os
import html
import pandas as pd
import schedule
import time
import pickle
import smtplib
import locale
import shutil
import tkinter as tk
import threading
from tkinter import scrolledtext
from datetime import datetime
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Rutas de archivos y variables necesarias
user_input = None
lock = threading.Lock()
global t
t = None
locale.setlocale(locale.LC_TIME, 'es_ES')
current_dir = os.getcwd()
csv_oc_pendientes = os.path.join(current_dir, 'OCs_Pendientes.csv')
pickle_file = os.path.join(current_dir, 'OCS\\facturas_procesadas.pickle')
ruta_excel_salida = os.path.join(current_dir, 'Facturas Sierra.xlsx')
ruta_terceros_csv = os.path.join(current_dir, 'terceros.csv')

# Funciones principales
def ha_cambiado():
    carpeta_backup = os.path.join(current_dir, 'backups')
    if not os.path.exists(carpeta_backup):
        os.makedirs(carpeta_backup, exist_ok=True)
        return True  # Si no existe la carpeta, asumimos que necesitamos hacer un backup

    backups_pickles = sorted([f for f in os.listdir(carpeta_backup) if f.endswith('.pickle')])
    if not backups_pickles:
        return True  # Si no hay backups, asumimos que necesitamos hacer uno

    ultimo_backup = os.path.join(carpeta_backup, backups_pickles[-1])
    try:
        with open(pickle_file, 'rb') as f_actual, open(ultimo_backup, 'rb') as f_ultimo:
            datos_actuales = pickle.load(f_actual)
            datos_ultimo_backup = pickle.load(f_ultimo)
    except FileNotFoundError:
        return True  # Si alguno de los archivos no existe, asumimos que hay un cambio

    return datos_actuales != datos_ultimo_backup

def guardar_backup_si_ha_cambiado():
    if ha_cambiado():  # Llamada sin argumentos
        fecha_actual = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        carpeta_backup = os.path.join(current_dir, 'backups')
        os.makedirs(carpeta_backup, exist_ok=True)
        
        # Definir las rutas de los archivos de backup
        archivo_backup_pickle = os.path.join(carpeta_backup, f'facturas_procesadas_{fecha_actual}.pickle')
        archivo_backup_excel = os.path.join(carpeta_backup, f'salida_{fecha_actual}.xlsx')
        archivo_backup_csv_oc_pendientes = os.path.join(carpeta_backup, f'OCs_Pendientes_{fecha_actual}.csv')
        
        # Copiar los archivos a la carpeta de backups
        shutil.copy(pickle_file, archivo_backup_pickle)
        shutil.copy(ruta_excel_salida, archivo_backup_excel)
        if os.path.exists(csv_oc_pendientes):
            shutil.copy(csv_oc_pendientes, archivo_backup_csv_oc_pendientes)
        
        # Mantenimiento de la cantidad de backups
        backups = sorted([f for f in os.listdir(carpeta_backup) if f.endswith('.pickle') or f.endswith('.xlsx') or f.endswith('.csv')], reverse=True)
        while len(backups) > 300:  # Asumiendo 100 versiones de cada tipo de archivo
            os.remove(os.path.join(carpeta_backup, backups.pop()))

        print("Backup realizado con éxito.")

def enviar_correo(asunto, cuerpo, destinatario, adjuntos=[]):
    emisor = 'facturas_gpf_sierra@outlook.com'  # Dirección de correo electrónico del emisor.
    contraseña = 'cnvzpbgggmtdqiry' #Clave de API correo

    mensaje = MIMEMultipart()  # Crea un objeto MIMEMultipart para el mensaje.
    mensaje['From'] = emisor  # Establece el emisor.
    mensaje['To'] = destinatario  # Establece el destinatario.
    mensaje['Subject'] = asunto  # Establece el asunto del correo.

    mensaje.attach(MIMEText(cuerpo, 'plain'))  # Adjunta el cuerpo del mensaje como texto plano.

    for archivo in adjuntos:  # Procesa cada archivo adjunto.
        parte = MIMEBase('application', 'octet-stream')  # Crea un objeto MIMEBase para el archivo.
        with open(archivo, 'rb') as file:  # Abre el archivo en modo binario.
            parte.set_payload(file.read())  # Lee y adjunta el contenido del archivo.
        encoders.encode_base64(parte)  # Codifica el contenido en base64.
        parte.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(archivo)}")  # Añade el nombre del archivo.
        mensaje.attach(parte)  # Adjunta el archivo al mensaje.

    #server = smtplib.SMTP('smtp.gmail.com', 587) # Conecta al servidor de Gmail usando SMTP en el puerto 587.
    server = smtplib.SMTP('smtp.office365.com', 587) ## Conecta al servidor de Outlook usando SMTP en el puerto 587.
    server.starttls()  # Inicia TLS para la seguridad de la conexión.
    server.login(emisor, contraseña)  # Inicia sesión con las credenciales del emisor.
    text = mensaje.as_string()  # Convierte el mensaje a una cadena de texto.
    server.sendmail(emisor, destinatario, text)  # Envía el correo.
    server.quit()  # Cierra la conexión con el servidor.

def cargar_o_inicializar_registros():
    try:
        with open(pickle_file, 'rb') as f:  # Intenta abrir el archivo pickle en modo lectura binaria.
            return pickle.load(f)  # Retorna el diccionario cargado desde el archivo pickle.
    except (FileNotFoundError, EOFError):  # Captura errores si el archivo no existe o está vacío.
        return {"facturas_procesadas": {}, "carpetas_vacias": {}}  # Retorna un nuevo diccionario con estructuras iniciales vacías.

def registrar_carpetas_vacias():
    registros = cargar_o_inicializar_registros()  # Carga o inicializa los registros.
    subcarpetas = [d for d in glob.glob(os.path.join(current_dir, 'OCS\\*\\'))]  # Lista todas las subcarpetas.
    
    for carpeta in subcarpetas:
        if not any(f.endswith('.xml') for f in os.listdir(carpeta)):  # Si no hay archivos .xml en la carpeta.
            oc = os.path.basename(carpeta.rstrip('\\'))  # Extrae el nombre de la OC basado en el nombre de la carpeta.
            registros["carpetas_vacias"][oc] = True  # Marca la OC como carpeta vacía en los registros.

    with open(pickle_file, 'wb') as f:  # Abre el archivo pickle en modo escritura binaria.
        pickle.dump(registros, f)  # Guarda los registros actualizados en el archivo pickle.

    # Actualiza el archivo CSV con las OCs pendientes basado en las carpetas vacías registradas.
    actualizar_csv_oc_pendientes(registros["carpetas_vacias"].keys())
    #Alerta
    print("registrando carpetas")

def actualizar_csv_oc_pendientes(ocs):
    df_oc_pendientes = pd.DataFrame({"OC": list(ocs)})  # Crea un DataFrame con las OCs pendientes.
    df_oc_pendientes.to_csv(csv_oc_pendientes, index=False)  # Guarda el DataFrame en un archivo CSV, sin índice.

def limpiar_registros_carpetas():
    registros = cargar_o_inicializar_registros()  # Carga los registros actuales.
    carpetas_a_eliminar = [carpeta for carpeta in registros["carpetas_vacias"] if any(f.endswith('.xml') for f in os.listdir(os.path.join(current_dir, 'OCS', carpeta)))] #Identifica carpetas a limpiar
                            
    for carpeta in carpetas_a_eliminar:  # Elimina las entradas de carpetas que ya no están vacías.
        del registros["carpetas_vacias"][carpeta]

    with open(pickle_file, 'wb') as f:  # Guarda los registros actualizados en el archivo pickle.
        pickle.dump(registros, f)

    actualizar_csv_oc_pendientes(registros["carpetas_vacias"].keys())  # Actualiza el archivo CSV de OCs pendientes.

def normalizar_ruc(ruc, longitud_estandar=13):
    # Asegura que el RUC tenga la longitud estándar, añadiendo ceros al inicio si es necesario
    return ruc.zfill(longitud_estandar)

def extraer_informacion_de_archivo(ruta_archivo):

    # Verificar si existe un archivo PDF con el mismo nombre que el archivo XML en la misma carpeta
    global user_input  # Declarar user_input como global
    ruta_pdf = os.path.splitext(ruta_archivo)[0] + '.pdf'
    nombre_carpeta = os.path.basename(os.path.dirname(ruta_archivo))

    # Iniciar un bucle que continúa hasta que el archivo PDF exista
    while not os.path.exists(ruta_pdf):
        print(f"Por favor verifique la OC en {nombre_carpeta}. Una vez solventado ingrese OK para continuar.")
        while True:
            with lock:
                if user_input is not None:
                    entrada = user_input.strip().lower()
                    print(f"Entrada recibida: {user_input}")  # Mostrar la entrada recibida
                    user_input = None  # Restablecer user_input para evitar repeticiones
                    
                    if entrada == "ok":
                        if os.path.exists(ruta_pdf):
                            print("Archivo PDF encontrado. Continuando con el proceso.")
                            break  # Salir del bucle si el archivo PDF ya existe
                        else:
                            print("El archivo PDF aún no existe. Por favor, verifique y vuelva a intentarlo.")
                    elif entrada == "":
                        print("No se detectó ninguna entrada. Por favor, ingrese OK cuando esté listo.")
                    else:
                        print("Entrada no reconocida. Ingrese OK para confirmar que el archivo PDF está listo.")
            time.sleep(1)  # Pequeña pausa para evitar saturación de CPU

    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as archivo:
            contenido = archivo.read()
    except UnicodeDecodeError:
        with open(ruta_archivo, 'r', encoding='ISO-8859-1', errors='replace') as archivo:
            contenido = archivo.read()

    contenido = html.unescape(contenido)

    patrones = {
        'ruc': r'<ruc>(.*?)<\/ruc>',
        'estab': r'<estab>(.*?)<\/estab>',
        'ptoEmi': r'<ptoEmi>(.*?)<\/ptoEmi>',
        'secuencial': r'<secuencial>(.*?)<\/secuencial>',
        'total_sin_impuestos': r'<totalSinImpuestos>(.*?)<\/totalSinImpuestos>',
        'fecha_emision': r'<fechaEmision>(.*?)<\/fechaEmision>',
        'nombre_comercial': r'<razonSocial>(.*?)<\/razonSocial>',
        'compania': r'<razonSocialComprador>(.*?)<\/razonSocialComprador>'
    }

    datos_extraidos = {}
    for clave, patron in patrones.items():
        coincidencia = re.search(patron, contenido, re.DOTALL)
        if coincidencia:
            datos_extraidos[clave] = coincidencia.group(1)
        else:
            datos_extraidos[clave] = 'No Disponible'

    if datos_extraidos['fecha_emision'] != 'No Disponible':
        datos_extraidos['fecha_formateada'] = datos_extraidos['fecha_emision']
    else:
        datos_extraidos['fecha_formateada'] = 'No Disponible'

    datos_extraidos['OC'] = os.path.basename(os.path.dirname(ruta_archivo))

    ruc = datos_extraidos.get('ruc')
    ruc_normalizado = normalizar_ruc(ruc)
    mapeo_terceros = cargar_y_mapear_terceros(ruta_terceros_csv)
    datos_extraidos['Tercero'] = mapeo_terceros.get(ruc_normalizado, {}).get('TERCERO', 'No Disponible')
    datos_extraidos['Centro de Costo'] = mapeo_terceros.get(ruc_normalizado, {}).get('CC', 'No Disponible')
    datos_extraidos['Nombre Farmacia'] = mapeo_terceros.get(ruc_normalizado, {}).get('NOMBRE FARMACIA', 'No Disponible')
    datos_extraidos['Frecuencia facturación'] = mapeo_terceros.get(ruc_normalizado, {}).get('FACTURA SEMESTRAL/MENSUAL', 'No Disponible')

    descripcion_tags = re.findall(r'<descripcion>(.*?)<\/descripcion>', contenido)
    precio_unitario_tags = re.findall(r'<precioUnitario>(.*?)<\/precioUnitario>', contenido)
    cantidad_tags = re.findall(r'<cantidad>(.*?)<\/cantidad>', contenido)

    precio_unitario_redondeado = [f"{float(precio):.2f}" for precio in precio_unitario_tags]
    cantidad_redondeado = [f"{float(cantidad):.2f}" for cantidad in cantidad_tags]

    descriptions_with_prices = [
        f"{descripcion} ({precio_unitario} x {cantidad})"
        for descripcion, precio_unitario, cantidad in zip(descripcion_tags, precio_unitario_redondeado, cantidad_redondeado)
    ]
    datos_extraidos['descripciones'] = " - ".join(descriptions_with_prices)
    
    return datos_extraidos

def actualizar_tabla_excel_y_limpieza(ruta_excel_salida):
    # Verifica si el archivo existe, si no, crea un archivo vacío con una hoja inicial
    inicializar = not os.path.exists(ruta_excel_salida)
    if inicializar:
        with pd.ExcelWriter(ruta_excel_salida, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name='Hoja_Temporal', index=False)  # Crea una hoja temporal vacía

    # Asumiendo que `current_dir`, `pickle_file`, y `csv_oc_pendientes` están definidos en el ámbito global o importados previamente
    archivos = glob.glob(os.path.join(current_dir, 'OCS', '**', '*.xml'), recursive=True)
    dataframe_total = pd.DataFrame()

    with open(pickle_file, 'rb') as f:
        facturas_procesadas = pickle.load(f)
    df_oc_pendientes = pd.read_csv(csv_oc_pendientes)

    for ruta_archivo in archivos:
        informacion = extraer_informacion_de_archivo(ruta_archivo)
        factura = f"{informacion['estab']}-{informacion['ptoEmi']}-{informacion['secuencial']}"
        oc = informacion['OC']

        if factura not in facturas_procesadas:
            facturas_procesadas[factura] = True
            descripcion = informacion['descripciones']
            dataframe_temporal = pd.DataFrame({
                'RUC': [informacion['ruc']],
                'Tercero': [informacion['Tercero']],
                'Nombre Comercial': [informacion['nombre_comercial']],
                'Compañía': [informacion['compania']],
                'Centro de Costo': [informacion['Centro de Costo']],
                'Nombre Farmacia': [informacion['Nombre Farmacia']],
                'Factura': [factura],
                'Total Sin Impuestos': [informacion['total_sin_impuestos']],
                'Fecha': [informacion['fecha_formateada']],
                'OC': [oc],
                'Frecuencia facturación': [informacion['Frecuencia facturación']],
                'Descripcion': [descripcion],
            })

            dataframe_temporal['Fecha de Envío Correo'] = pd.to_datetime('today').strftime('%Y-%m-%d')
            dataframe_total = pd.concat([dataframe_total, dataframe_temporal], ignore_index=True)

            if oc in facturas_procesadas and not facturas_procesadas[oc]:
                facturas_procesadas[oc] = True
                df_oc_pendientes = df_oc_pendientes[df_oc_pendientes['OC'] != oc]

            asunto = f"FACTURA ARRIENDO {informacion['compania']} No {factura}"
            cuerpo = f"Buen día estimados, \n Por favor su gentil ayuda con el registro de la factura \n Factura No: {factura} \n OC: {oc}"
            destinatario = 'g_gyerecepcionfacturasservicios@corporaciongpf.com'
            ruta_xml = ruta_archivo
            ruta_pdf = ruta_archivo.replace('.xml', '.pdf')
            enviar_correo(asunto, cuerpo, destinatario, [ruta_xml, ruta_pdf])
            print(f"Enviando correo OC {oc}")


    if not dataframe_total.empty:
        dataframe_total['Fecha_convertida'] = pd.to_datetime(dataframe_total['Fecha'], format='%d/%m/%Y', errors='coerce')
        meses = dataframe_total['Fecha_convertida'].dt.strftime('%B %Y').unique()

        for mes in meses:
            df_mes = dataframe_total[dataframe_total['Fecha_convertida'].dt.strftime('%B %Y') == mes]

            with pd.ExcelWriter(ruta_excel_salida, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if mes in writer.book.sheetnames:
                    startrow = writer.sheets[mes].max_row
                else:
                    startrow = 0
                df_mes.to_excel(writer, sheet_name=mes, index=False, header=not bool(startrow), startrow=startrow)

        # Eliminar la hoja temporal si se inicializó el archivo
        if inicializar:
            wb = load_workbook(ruta_excel_salida)
            if 'Hoja_Temporal' in wb.sheetnames:
                del wb['Hoja_Temporal']
            wb.save(ruta_excel_salida)

    with open(pickle_file, 'wb') as f:
        pickle.dump(facturas_procesadas, f)
    df_oc_pendientes.to_csv(csv_oc_pendientes, index=False)

    print("Archivo Excel Actualizado")

def cargar_y_mapear_terceros(ruta_terceros_csv):
    # Intenta leer el archivo CSV con diferentes codecs
    try:
        terceros_df = pd.read_csv(ruta_terceros_csv, encoding='utf-8')
    except UnicodeDecodeError:
        terceros_df = pd.read_csv(ruta_terceros_csv, encoding='latin1')  # Prueba con el codec latin1

    terceros_df['RUC'] = terceros_df['RUC'].apply(lambda x: normalizar_ruc(str(x)))
    #mapeo_terceros = terceros_df.set_index('RUC')['TERCERO'].to_dict()
    terceros_df.drop_duplicates(subset='RUC', inplace=True)
    mapeo_terceros = terceros_df.set_index('RUC')[['TERCERO', 'CC', 'NOMBRE FARMACIA', 'FACTURA SEMESTRAL/MENSUAL']].to_dict(orient='index')


    return mapeo_terceros

def main():
    # Programa las otras tareas para ejecución periódica
    tarea1 = schedule.every(10).seconds.do(registrar_carpetas_vacias)
    tarea2 = schedule.every(10).seconds.do(limpiar_registros_carpetas)
    tarea3 = schedule.every(10).seconds.do(actualizar_tabla_excel_y_limpieza, ruta_excel_salida)
    tarea4 = schedule.every(10).seconds.do(guardar_backup_si_ha_cambiado)

    try:
        while True:
            if stop_thread:  # Verifica si se ha señalado la detención
                print("Deteniendo tareas programadas...")
                # Cancela todas las tareas programadas
                schedule.cancel_job(tarea1)
                schedule.cancel_job(tarea2)
                schedule.cancel_job(tarea3)
                schedule.cancel_job(tarea4)
                break  # Sale del bucle
            schedule.run_pending()  # Ejecuta las tareas pendientes según su programación.
            time.sleep(1)  # Espera 1 segundo antes de la próxima verificación de tareas pendientes.
    except Exception as e:
        print(f"Error durante la ejecución de tareas: {e}")
    finally:
        print("Finalizando el programa principal.")

def iniciar_tareas():
    global t, stop_thread
    print("Iniciando gestión de facturas, por favor espere.")
    stop_thread = False
    t = threading.Thread(target=ejecutar_tareas)
    t.start()

def check_thread():
    global t
    if t.is_alive():
        # Si el hilo todavía está corriendo, revisa de nuevo en 100 ms
        window.after(100, check_thread)
    else:
        # Cuando el hilo termina, actualiza la interfaz como necesites
        print("Todas las tareas han sido detenidas.")
        stop_button.config(relief=tk.RAISED)  # Cambia el relieve del botón si está presionado

def detener_tareas():
    global t, stop_thread
    stop_thread = True
    print("Deteniendo todas las tareas, por favor espere...")
    stop_button.config(relief=tk.SUNKEN)  # Hace que el botón parezca presionado
    check_thread()  # Comienza a verificar si el hilo ha terminado

def ejecutar_tareas():
    try:
        while not stop_thread:
            main()
            time.sleep(1)
            pass
    except KeyboardInterrupt:
        print("Tareas detenidas.")

def enviar_input():
    global user_input
    entrada = entry_box.get()
    entry_box.delete(0, tk.END)
    with lock:
        user_input = entrada
    print(f"Entrada recibida: {entrada}")

# Configuración de la ventana principal
window = tk.Tk()
window.title("Gestión de Facturas GPF")

# Área de texto para salida
text_area = scrolledtext.ScrolledText(window, wrap=tk.WORD, width=40, height=10)
text_area.grid(column=0, row=0, columnspan=3, pady=10, padx=10)

# Redefinir la función print para que muestre en el área de texto
def print(*args, **kwargs):
    text_area.insert(tk.END, ' '.join(map(str, args)) + '\n')
    text_area.see(tk.END)

# Entrada de texto
entry_box = tk.Entry(window, width=25)
entry_box.grid(column=0, row=1, pady=10)

# Botón para enviar input
input_button = tk.Button(window, text="Enviar", command=enviar_input)
input_button.grid(column=1, row=1)

# Botones de control
start_button = tk.Button(window, text="Iniciar", command=iniciar_tareas)
start_button.grid(column=0, row=2, pady=10)

stop_button = tk.Button(window, text="Detener", command=detener_tareas)
stop_button.grid(column=1, row=2)

window.mainloop()