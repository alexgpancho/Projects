# V1.02 Alexis G

# pip install pyinstaller xlsxwriter pandas openpyxl cryptography smartsheet-python-sdk azure-identity httpx
# dependencias

# Importar librerías
import re
import glob
import os  
import html
import pandas as pd
import time
import pickle
import locale
import shutil
import tkinter as tk
import threading
from tkinter import scrolledtext
from datetime import datetime
from openpyxl import load_workbook
import asyncio
#Librerías locales
from security import security
from manejo_correo import enviar_correo, eliminar_correos_enviados, autenticar

# Rutas de archivos y variables necesarias
user_input = None
lock = threading.Lock()
global t
t = None
locale.setlocale(locale.LC_TIME, 'es_ES')
current_dir = os.getcwd()
csv_oc_pendientes = os.path.join(current_dir, 'OCs_Pendientes.csv')
pickle_file = os.path.join(current_dir, 'OCS\\facturas_procesadas.pickle')
ruta_excel_salida = os.path.join(current_dir, 'Facturas Sierra.xlsx')
ruta_terceros_csv = os.path.join(current_dir, 'terceros.csv')
ruta_destinatarios = os.path.join(current_dir,'destinatarios.csv')

# Funciones principales
def ha_cambiado():
    carpeta_backup = os.path.join(current_dir, 'backups')
    if not os.path.exists(carpeta_backup):
        os.makedirs(carpeta_backup, exist_ok=True)
        return True  # Si no existe la carpeta, asumimos que necesitamos hacer un backup

    backups_pickles = sorted([f for f in os.listdir(carpeta_backup) if f.endswith('.pickle')])
    if not backups_pickles:
        return True  # Si no hay backups, asumimos que necesitamos hacer uno

    ultimo_backup = os.path.join(carpeta_backup, backups_pickles[-1])
    try:
        with open(pickle_file, 'rb') as f_actual, open(ultimo_backup, 'rb') as f_ultimo:
            datos_actuales = pickle.load(f_actual)
            datos_ultimo_backup = pickle.load(f_ultimo)
    except FileNotFoundError:
        return True  # Si alguno de los archivos no existe, asumimos que hay un cambio

    return datos_actuales != datos_ultimo_backup

def guardar_backup_si_ha_cambiado():
    if ha_cambiado():  # Llamada sin argumentos
        fecha_actual = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        carpeta_backup = os.path.join(current_dir, 'backups')
        os.makedirs(carpeta_backup, exist_ok=True)
        
        # Definir las rutas de los archivos de backup
        archivo_backup_pickle = os.path.join(carpeta_backup, f'facturas_procesadas_{fecha_actual}.pickle')
        archivo_backup_excel = os.path.join(carpeta_backup, f'salida_{fecha_actual}.xlsx')
        archivo_backup_csv_oc_pendientes = os.path.join(carpeta_backup, f'OCs_Pendientes_{fecha_actual}.csv')
        
        # Copiar los archivos a la carpeta de backups
        shutil.copy(pickle_file, archivo_backup_pickle)
        shutil.copy(ruta_excel_salida, archivo_backup_excel)
        if os.path.exists(csv_oc_pendientes):
            shutil.copy(csv_oc_pendientes, archivo_backup_csv_oc_pendientes)
        
        # Mantenimiento de la cantidad de backups
        backups = sorted([f for f in os.listdir(carpeta_backup) if f.endswith('.pickle') or f.endswith('.xlsx') or f.endswith('.csv')], reverse=True)
        while len(backups) > 20:  # Asumiendo 20 versiones de cada tipo de archivo
            os.remove(os.path.join(carpeta_backup, backups.pop()))

        print("Backup realizado con éxito.")

def cargar_o_inicializar_registros():
    if not os.path.exists(pickle_file):
        with open(pickle_file, 'wb') as f:  # Crea un archivo pickle vacío si no existe.
            pickle.dump({"facturas_procesadas": {}, "carpetas_vacias": {}}, f)
    
    try:
        with open(pickle_file, 'rb') as f:  # Intenta abrir el archivo pickle en modo lectura binaria.
            return pickle.load(f)  # Retorna el diccionario cargado desde el archivo pickle.
    except (FileNotFoundError, EOFError):  # Captura errores si el archivo no existe o está vacío.
        return {"facturas_procesadas": {}, "carpetas_vacias": {}}  # Retorna un nuevo diccionario con estructuras iniciales vacías.


def registrar_carpetas_vacias():
    registros = cargar_o_inicializar_registros()  # Carga o inicializa los registros.
    subcarpetas = [d for d in glob.glob(os.path.join(current_dir, 'OCS\\*\\'))]  # Lista todas las subcarpetas.
    
    for carpeta in subcarpetas:
        if not any(f.endswith('.xml') for f in os.listdir(carpeta)):  # Si no hay archivos .xml en la carpeta.
            oc = os.path.basename(carpeta.rstrip('\\'))  # Extrae el nombre de la OC basado en el nombre de la carpeta.
            registros["carpetas_vacias"][oc] = True  # Marca la OC como carpeta vacía en los registros.

    if not os.path.exists(pickle_file):
        with open(pickle_file, 'wb') as f:  # Crea un archivo pickle vacío si no existe.
            pickle.dump({"facturas_procesadas": {}, "carpetas_vacias": {}}, f)

    # Actualiza el archivo CSV con las OCs pendientes basado en las carpetas vacías registradas.
    actualizar_csv_oc_pendientes(registros["carpetas_vacias"].keys())
    #Alerta
    print("Registrando carpetas")

def actualizar_csv_oc_pendientes(ocs):
    df_oc_pendientes = pd.DataFrame({"OC": list(ocs)})  # Crea un DataFrame con las OCs pendientes.
    df_oc_pendientes.to_csv(csv_oc_pendientes, index=False)  # Guarda el DataFrame en un archivo CSV, sin índice.

def limpiar_registros_carpetas():
    registros = cargar_o_inicializar_registros()  # Carga los registros actuales.
    carpetas_a_eliminar = [carpeta for carpeta in registros["carpetas_vacias"] if any(f.endswith('.xml') for f in os.listdir(os.path.join(current_dir, 'OCS', carpeta)))] #Identifica carpetas a limpiar

    for carpeta in carpetas_a_eliminar:  # Elimina las entradas de carpetas que ya no están vacías.
        del registros["carpetas_vacias"][carpeta]

    with open(pickle_file, 'wb') as f:  # Guarda los registros actualizados en el archivo pickle.
        pickle.dump(registros, f)

    actualizar_csv_oc_pendientes(registros["carpetas_vacias"].keys())  # Actualiza el archivo CSV de OCs pendientes.

def normalizar_ruc(ruc, longitud_estandar=13):
    # Asegura que el RUC tenga la longitud estándar, añadiendo ceros al inicio si es necesario
    return ruc.zfill(longitud_estandar)

def extraer_informacion_de_archivo(ruta_archivo):

    # Verificar si existe un archivo PDF con el mismo nombre que el archivo XML en la misma carpeta
    global user_input  # Declarar user_input como global
    ruta_pdf = os.path.splitext(ruta_archivo)[0] + '.pdf'
    nombre_carpeta = os.path.basename(os.path.dirname(ruta_archivo))

    # Iniciar un bucle que continúa hasta que el archivo PDF exista
    while not os.path.exists(ruta_pdf):
        print(f"Por favor verifique la OC en {nombre_carpeta}. Una vez solventado ingrese OK para continuar.")
        while True:
            with lock:
                if user_input is not None:
                    entrada = user_input.strip().lower()
                    print(f"Entrada recibida: {user_input}")  # Mostrar la entrada recibida
                    user_input = None  # Restablecer user_input para evitar repeticiones
                    
                    if entrada == "ok":
                        if os.path.exists(ruta_pdf):
                            print("Archivo PDF encontrado. Continuando con el proceso.")
                            break  # Salir del bucle si el archivo PDF ya existe
                        else:
                            print("El archivo PDF aún no existe. Por favor, verifique y vuelva a intentarlo.")
                    elif entrada == "":
                        print("No se detectó ninguna entrada. Por favor, ingrese OK cuando esté listo.")
                    else:
                        print("Entrada no reconocida. Ingrese OK para confirmar que el archivo PDF está listo.")
            time.sleep(1)  # Pequeña pausa para evitar saturación de CPU

    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as archivo:
            contenido = archivo.read()
    except UnicodeDecodeError:
        with open(ruta_archivo, 'r', encoding='ISO-8859-1', errors='replace') as archivo:
            contenido = archivo.read()

    contenido = html.unescape(contenido)

    patrones = {
        'autorizacion': r'<numeroAutorizacion>(.*?)<\/numeroAutorizacion>',
        'ruc': r'<ruc>(.*?)<\/ruc>',
        'estab': r'<estab>(.*?)<\/estab>',
        'ptoEmi': r'<ptoEmi>(.*?)<\/ptoEmi>',
        'secuencial': r'<secuencial>(.*?)<\/secuencial>',
        'fecha_emision': r'<fechaEmision>(.*?)<\/fechaEmision>',
        'nombre_comercial': r'<razonSocial>(.*?)<\/razonSocial>',
        'compania': r'<razonSocialComprador>(.*?)<\/razonSocialComprador>'
    }

    datos_extraidos = {}
    for clave, patron in patrones.items():
        coincidencia = re.search(patron, contenido, re.DOTALL)
        if coincidencia:
            datos_extraidos[clave] = coincidencia.group(1)
        else:
            datos_extraidos[clave] = 'No Disponible'

    if datos_extraidos['fecha_emision'] != 'No Disponible':
        datos_extraidos['fecha_formateada'] = datos_extraidos['fecha_emision']
    else:
        datos_extraidos['fecha_formateada'] = 'No Disponible'

    datos_extraidos['OC'] = os.path.basename(os.path.dirname(ruta_archivo))

    ruc = datos_extraidos.get('ruc')
    ruc_normalizado = normalizar_ruc(ruc)
    mapeo_terceros = cargar_y_mapear_terceros(ruta_terceros_csv)
    datos_extraidos['Tercero'] = mapeo_terceros.get(ruc_normalizado, {}).get('TERCERO', 'No Disponible')
    datos_extraidos['Centro de Costo'] = mapeo_terceros.get(ruc_normalizado, {}).get('CC', 'No Disponible')
    datos_extraidos['Nombre Farmacia'] = mapeo_terceros.get(ruc_normalizado, {}).get('NOMBRE FARMACIA', 'No Disponible')
    datos_extraidos['Frecuencia facturación'] = mapeo_terceros.get(ruc_normalizado, {}).get('FACTURA SEMESTRAL/MENSUAL', 'No Disponible')

    descripcion_tags = re.findall(r'<descripcion>(.*?)<\/descripcion>', contenido)
    precio_unitario_tags = re.findall(r'<precioUnitario>(.*?)<\/precioUnitario>', contenido)
    cantidad_tags = re.findall(r'<cantidad>(.*?)<\/cantidad>', contenido)

    precio_unitario_redondeado = [f"{float(precio):.2f}" for precio in precio_unitario_tags]
    cantidad_redondeado = [f"{float(cantidad):.2f}" for cantidad in cantidad_tags]

    descriptions_with_prices = [
        f"{descripcion} ({precio_unitario} x {cantidad})"
        for descripcion, precio_unitario, cantidad in zip(descripcion_tags, precio_unitario_redondeado, cantidad_redondeado)
    ]
    datos_extraidos['descripciones'] = " - ".join(descriptions_with_prices)
    
    subtotales = {}
    iva_values = {}

    # Buscar en la etiqueta "detalle" -> "impuestos" -> "impuesto"
    detalle_pattern = r'<detalle>(.*?)<\/detalle>'
    detalle_matches = re.findall(detalle_pattern, contenido, re.DOTALL)

    for detalle in detalle_matches:
        impuestos_pattern = r'<impuesto>(.*?)<\/impuesto>'
        impuestos_matches = re.findall(impuestos_pattern, detalle, re.DOTALL)

        for impuesto in impuestos_matches:
            base_imponible_pattern = r'<baseImponible>(.*?)<\/baseImponible>'
            tarifa_pattern = r'<tarifa>(.*?)<\/tarifa>'
            valor_pattern = r'<valor>(.*?)<\/valor>'

            base_imponible = re.search(base_imponible_pattern, impuesto)
            tarifa = re.search(tarifa_pattern, impuesto)
            valor = re.search(valor_pattern, impuesto)

            if base_imponible and tarifa and valor:
                base_imponible = float(base_imponible.group(1))
                tarifa = float(tarifa.group(1))
                valor = float(valor.group(1))

                if tarifa not in subtotales:
                    subtotales[tarifa] = 0
                if tarifa not in iva_values:
                    iva_values[tarifa] = 0

                subtotales[tarifa] += base_imponible
                iva_values[tarifa] += valor

    if not subtotales and not iva_values:
        total_impuestos_pattern = r'<totalImpuesto>(.*?)<\/totalImpuesto>'
        total_impuestos_matches = re.findall(total_impuestos_pattern, contenido, re.DOTALL)

        for total_impuesto in total_impuestos_matches:
            base_imponible_pattern = r'<baseImponible>(.*?)<\/baseImponible>'
            tarifa_pattern = r'<tarifa>(.*?)<\/tarifa>'
            valor_pattern = r'<valor>(.*?)<\/valor>'

            base_imponible = re.search(base_imponible_pattern, total_impuesto)
            tarifa = re.search(tarifa_pattern, total_impuesto)
            valor = re.search(valor_pattern, total_impuesto)

            if base_imponible and tarifa and valor:
                base_imponible = float(base_imponible.group(1))
                tarifa = float(tarifa.group(1))
                valor = float(valor.group(1))

                if tarifa not in subtotales:
                    subtotales[tarifa] = 0
                if tarifa not in iva_values:
                    iva_values[tarifa] = 0

                subtotales[tarifa] += base_imponible
                iva_values[tarifa] += valor

    return datos_extraidos, subtotales, iva_values

def actualizar_tabla_excel_y_limpieza(ruta_excel_salida, access_token):
    # Verifica si el archivo existe, si no, crea un archivo vacío con una hoja inicial
    inicializar = not os.path.exists(ruta_excel_salida)
    if inicializar:
        with pd.ExcelWriter(ruta_excel_salida, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name='Hoja_Temporal', index=False)  # Crea una hoja temporal vacía

    # Asumiendo que `current_dir`, `pickle_file`, y `csv_oc_pendientes` están definidos en el ámbito global o importados previamente
    archivos = glob.glob(os.path.join(current_dir, 'OCS', '**', '*.xml'), recursive=True)
    dataframe_total = pd.DataFrame()

    with open(pickle_file, 'rb') as f:
        facturas_procesadas = pickle.load(f)
    df_oc_pendientes = pd.read_csv(csv_oc_pendientes)

    # Intentar leer el archivo CSV con diferentes codificaciones
    try:
        df = pd.read_csv(ruta_destinatarios, encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(ruta_destinatarios, encoding='latin1')
    
    destinatario = df['destinatario'].iloc[0]
    cc = df['cc'].iloc[0]

    for ruta_archivo in archivos:
        informacion, subtotales, iva_values = extraer_informacion_de_archivo(ruta_archivo)
        factura = f"{informacion['estab']}-{informacion['ptoEmi']}-{informacion['secuencial']}"
        oc = informacion['OC']

        if oc not in facturas_procesadas:
            facturas_procesadas[oc] = True
            descripcion = informacion['descripciones']

            # Preparar las nuevas columnas
            subtotal_0 = subtotales.get(0, 0)
            tarifas_no_0 = ";".join([str(tarifa) for tarifa in subtotales if tarifa != 0])
            subtotales_impuesto_no_0 = ";".join([str(subtotales[tarifa]) for tarifa in subtotales if tarifa != 0])
            iva_no_0 = ";".join([str(iva_values[tarifa]) for tarifa in iva_values if tarifa != 0])

            dataframe_temporal = pd.DataFrame({
                'Autorizacion': [informacion["autorizacion"]],
                'RUC': [informacion['ruc']],
                'Tercero': [informacion['Tercero']],
                'Nombre Comercial': [informacion['nombre_comercial']],
                'Compañía': [informacion['compania']],
                'Centro de Costo': [informacion['Centro de Costo']],
                'Nombre Farmacia': [informacion['Nombre Farmacia']],
                'OC': [oc],
                'Factura': [factura],
                'Fecha': [informacion['fecha_formateada']],
                'Descripcion': [descripcion],
                'Subtotal 0%': [subtotal_0],
                'Tarifa': [tarifas_no_0],
                'Subtotales Impuesto': [subtotales_impuesto_no_0],
                'IVA': [iva_no_0],
                'Frecuencia facturación': [informacion['Frecuencia facturación']]
            })

            dataframe_temporal['Fecha de Envío Correo'] = pd.to_datetime('today').strftime('%Y-%m-%d')
            dataframe_total = pd.concat([dataframe_total, dataframe_temporal], ignore_index=True)

            if oc in facturas_procesadas and not facturas_procesadas[oc]:
                facturas_procesadas[oc] = True
                df_oc_pendientes = df_oc_pendientes[df_oc_pendientes['OC'] != oc]

            asunto = f"FACTURA ARRIENDO {informacion['compania']} No {factura}"
            cuerpo = f"Buen día estimados, \n Por favor su gentil ayuda con el registro de la factura \n Factura No: {factura} \n OC: {oc}"
            ruta_xml = ruta_archivo
            ruta_pdf = ruta_archivo.replace('.xml', '.pdf')
            asyncio.run(enviar_correo(asunto, cuerpo, destinatario, cc, [ruta_xml, ruta_pdf], access_token, print)) #OJO
            print(f"OC Nro: {oc}")

    if not dataframe_total.empty:
        dataframe_total['Fecha_convertida'] = pd.to_datetime(dataframe_total['Fecha'], format='%d/%m/%Y', errors='coerce')
        meses = dataframe_total['Fecha_convertida'].dt.strftime('%B %Y').unique()

        for mes in meses:
            df_mes = dataframe_total[dataframe_total['Fecha_convertida'].dt.strftime('%B %Y') == mes]

            # Eliminar la columna 'Fecha_convertida' antes de escribir en el archivo
            df_mes = df_mes.drop(columns=['Fecha_convertida'])

            with pd.ExcelWriter(ruta_excel_salida, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if mes in writer.book.sheetnames:
                    startrow = writer.sheets[mes].max_row
                else:
                    startrow = 0
                df_mes.to_excel(writer, sheet_name=mes, index=False, header=not bool(startrow), startrow=startrow)
                writer.book.close()

        # Eliminar la hoja temporal si se inicializó el archivo
        if inicializar:
            wb = load_workbook(ruta_excel_salida)
            if 'Hoja_Temporal' in wb.sheetnames:
                del wb['Hoja_Temporal']
            wb.save(ruta_excel_salida)
            wb.close()

    with open(pickle_file, 'wb') as f:
        pickle.dump(facturas_procesadas, f)
    df_oc_pendientes.to_csv(csv_oc_pendientes, index=False)


def cargar_y_mapear_terceros(ruta_terceros_csv):
    # Intenta leer el archivo CSV con diferentes codecs
    try:
        terceros_df = pd.read_csv(ruta_terceros_csv, encoding='utf-8')
    except UnicodeDecodeError:
        terceros_df = pd.read_csv(ruta_terceros_csv, encoding='latin1')  # Prueba con el codec latin1

    terceros_df['RUC'] = terceros_df['RUC'].apply(lambda x: normalizar_ruc(str(x)))
    terceros_df.drop_duplicates(subset='RUC', inplace=True)
    mapeo_terceros = terceros_df.set_index('RUC')[['TERCERO', 'CC', 'NOMBRE FARMACIA', 'FACTURA SEMESTRAL/MENSUAL']].to_dict(orient='index')


    return mapeo_terceros

def main():
    try:
        access_token = asyncio.run(autenticar(print)) #OJO
        registrar_carpetas_vacias()
        limpiar_registros_carpetas()
        actualizar_tabla_excel_y_limpieza(ruta_excel_salida, access_token)
        time.sleep(5) #OJO
        asyncio.run(eliminar_correos_enviados(print, access_token)) #OJO
        print("Archivo Excel Actualizado")
        guardar_backup_si_ha_cambiado() #OJO
    except Exception as e:
        print(f"Error durante la ejecución de tareas: {e}")
    finally:
        print("Finalizando la ejecución de tareas.")

# Código interfaz gráfica
def iniciar_tareas():
    print("Iniciando gestión de facturas, por favor espere.")
    t = threading.Thread(target=main)
    t.start()

def enviar_input():
    global user_input
    entrada = entry_box.get()
    entry_box.delete(0, tk.END)
    user_input = entrada
    print(f"Entrada recibida")
    if validar_clave(entrada):
        print("Clave válida. Puede iniciar las tareas.")
        start_button.config(state=tk.NORMAL)
    else:
        print("Clave inválida. Inténtelo de nuevo.")
        start_button.config(state=tk.DISABLED)

def validar_clave(clave_ingresada):
    datos = security()
    return clave_ingresada == datos['clave']

window = tk.Tk()
window.title("Gestión de Facturas GPF")

text_area = scrolledtext.ScrolledText(window, wrap=tk.WORD, width=45, height=14)
text_area.grid(column=0, row=0, columnspan=3, pady=10, padx=10, sticky="nsew")

window.grid_columnconfigure(0, weight=1)
window.grid_columnconfigure(1, weight=1)
window.grid_columnconfigure(2, weight=1)
window.grid_rowconfigure(0, weight=1)

def print(*args, **kwargs):
    text_area.insert(tk.END, ' '.join(map(str, args)) + '\n')
    text_area.see(tk.END)

print("Bienvenido. Por favor, ingrese la clave para iniciar.")

entry_box = tk.Entry(window, width=35)
entry_box.grid(column=0, row=1, pady=10)

input_button = tk.Button(window, text="Enviar", command=enviar_input)
input_button.grid(column=1, row=1)

start_button = tk.Button(window, text="Iniciar", command=iniciar_tareas)
start_button.grid(column=0, row=2, pady=10)
start_button.config(state=tk.DISABLED)

stop_button = tk.Button(window, text="Terminar", command=lambda: os._exit(0))
stop_button.grid(column=1, row=2)

window.mainloop()